from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory
from werkzeug.utils import secure_filename
import os
import docx
import PyPDF2
import openpyxl
import subprocess
import re

app = Flask(__name__)
app.secret_key = "supersecretkey"  # change in production

# Allowed file types
ALLOWED_EXTENSIONS = {"txt", "pdf", "docx", "xlsx"}

# Upload folders
UPLOAD_FOLDER = "uploads"
TEXT_FOLDER = os.path.join(UPLOAD_FOLDER, "text-extraction")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TEXT_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["TEXT_FOLDER"] = TEXT_FOLDER


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_text(filepath, ext):
    """Extracts plain text from supported file formats."""
    text = ""
    try:
        if ext == "txt":
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()

        elif ext == "pdf":
            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() or ""

        elif ext == "docx":
            doc = docx.Document(filepath)
            for para in doc.paragraphs:
                text += para.text + "\n"

        elif ext == "xlsx":
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"

    except Exception as e:
        text = f"Error extracting text: {e}"

    return text


@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        if username == "admin" and password == "admin":
            session["user"] = username
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid login!", "danger")
    return render_template("login.html")


@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        if "file" not in request.files:
            flash("No file part", "danger")
            return redirect(request.url)

        files = request.files.getlist("file")
        if not files:
            flash("No selected files", "danger")
            return redirect(request.url)

        uploaded_count = 0
        skipped_count = 0

        for file in files:
            if not file or file.filename == "":
                continue

            filename = secure_filename(file.filename)

            if file and allowed_file(filename):
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                # Extract text (backend only)
                ext = filename.rsplit(".", 1)[1].lower()
                text_content = extract_text(filepath, ext)
                text_filename = os.path.splitext(filename)[0] + ".txt"
                text_path = os.path.join(app.config["TEXT_FOLDER"], text_filename)

                with open(text_path, "w", encoding="utf-8") as f:
                    f.write(text_content)

                # 🔹 Auto-update pooled file after each extraction
                pool_extracted_files()

                uploaded_count += 1
            else:
                skipped_count += 1
                flash(f"Skipped invalid file: {filename}", "warning")

        if uploaded_count > 0:
            flash(f"Uploaded {uploaded_count} file(s) successfully!", "success")
        if skipped_count > 0:
            flash(f"Skipped {skipped_count} invalid file(s).", "danger")

    files = [
        f for f in os.listdir(app.config["UPLOAD_FOLDER"])
        if os.path.isfile(os.path.join(app.config["UPLOAD_FOLDER"], f))
    ]
    file_count = len(files)

    return render_template("dashboard.html", files=files, count=file_count)

@app.route("/chat", methods=["GET", "POST"])
def chat():
    if "user" not in session:
        return redirect(url_for("login"))

    conversation = session.get("conversation", [])

    if request.method == "POST":
        user_msg = request.form.get("message", "").strip()
        if user_msg:
            conversation.append(("You", user_msg))

            pooled_path = os.path.join(app.config["TEXT_FOLDER"], "pooled.txt")
            if os.path.exists(pooled_path):
                ai_reply = query_llama_bin(user_msg, pooled_path)
            else:
                ai_reply = "[AI placeholder] No pooled text available yet."

            conversation.append(("AI", ai_reply))
            session["conversation"] = conversation

    return render_template("chat.html", conversation=conversation)

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


def pool_extracted_files():
    pooled_path = os.path.join(app.config["TEXT_FOLDER"], "pooled.txt")
    seen_lines = set()  # keep track of unique lines

    with open(pooled_path, "w", encoding="utf-8") as pooled:
        for fname in os.listdir(app.config["TEXT_FOLDER"]):
            fpath = os.path.join(app.config["TEXT_FOLDER"], fname)
            if os.path.isfile(fpath) and fname.endswith(".txt") and fname != "pooled.txt":
                pooled.write(f"\n--- Start of {fname} ---\n")

                with open(fpath, "r", encoding="utf-8") as f:
                    for line in f:
                        clean_line = line.strip()
                        if clean_line and clean_line not in seen_lines:
                            pooled.write(clean_line + "\n")
                            seen_lines.add(clean_line)

                pooled.write(f"--- End of {fname} ---\n\n")

    return pooled_path

def truncate_context(context, max_chars=5000):
    """Keep only the last max_chars of text to fit model context"""
    return context[-max_chars:] if len(context) > max_chars else context

def strip_ansi(text):
    """Remove ANSI escape sequences"""
    ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
    return ansi_escape.sub('', text)

def query_llama_bin(question, pooled_file_path):
    # Load pooled.txt
    with open(pooled_file_path, "r", encoding="utf-8") as f:
        context = f.read()

    # Truncate context
    context = truncate_context(context, max_chars=4000)

    # Build prompt
    prompt = f"""You are an AI assistant.
Here is the knowledge base from pooled documents:
{context}

Now answer the following user question clearly and concisely:

Question: {question}
Answer:"""

    # Call llama-run.exe
    process = subprocess.Popen(
        [
            "llamacpp/llama-run.exe",
            "models/llama.gguf",
            prompt,
            "-t", "4",
            "-c", "5000",
            "--temp", "0.8",
            "--color", "false"
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True
    )

    stdout, stderr = process.communicate()

    if stderr:
        print("LLaMA Error:", stderr)

    return strip_ansi(stdout).strip()



if __name__ == "__main__":
    app.run(debug=True)
